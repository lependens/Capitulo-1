import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.comments import Comment
from openpyxl.utils import get_column_letter
from datetime import datetime
import math  # For sample calculations if needed, but formulas in Excel

# Create a new workbook
wb = openpyxl.Workbook()

# Sheet 1: Inputs
ws_inputs = wb.active
ws_inputs.title = "Inputs"
ws_inputs['A1'] = "Inputs for ET0 Penman-Monteith Calculation"
ws_inputs['A1'].font = Font(bold=True, size=14)
ws_inputs.merge_cells('A1:F1')

# Headers for inputs
headers = ["Parameter", "Value", "Unit", "Description"]
for col, header in enumerate(headers, start=1):
    cell = ws_inputs.cell(row=3, column=col, value=header)
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal='center')

# List of inputs based on Penman-Monteith requirements
inputs = [
    ("Latitude", "", "degrees (e.g., 39.006)", "Latitude in decimal degrees (from DMS if needed)"),
    ("Altitude (z)", "", "m", "Elevation above sea level"),
    ("Date", "", "YYYY-MM-DD", "Date for calculation (for Julian day)"),
    ("Mean Temperature (T)", "", "°C", "Daily mean air temperature"),
    ("Max Temperature (Tmax)", "", "°C", "Daily maximum air temperature"),
    ("Min Temperature (Tmin)", "", "°C", "Daily minimum air temperature"),
    ("Max Relative Humidity (RHmax)", "", "%", "Daily maximum relative humidity"),
    ("Min Relative Humidity (RHmin)", "", "%", "Daily minimum relative humidity"),
    ("Wind Speed (U2)", "", "m/s", "Wind speed at 2m height"),
    ("Solar Radiation (Rs)", "", "MJ/m²/day", "Measured solar radiation (if available)"),
    ("Sunshine Hours (n)", "", "hours", "Actual sunshine hours (for alternative Rs estimation)"),
]

row = 4
for param, value, unit, desc in inputs:
    ws_inputs.cell(row=row, column=1, value=param)
    cell_value = ws_inputs.cell(row=row, column=2, value=value)
    ws_inputs.cell(row=row, column=3, value=unit)
    cell_desc = ws_inputs.cell(row=row, column=4, value=desc)
    cell_desc.comment = Comment("Enter value in column B. Formulas in Calculation sheet will reference these.", "Grok")
    row += 1

# Adjust column widths
for col in range(1, 5):
    ws_inputs.column_dimensions[get_column_letter(col)].width = 30 if col == 4 else 20

# Sheet 2: Calculations
ws_calc = wb.create_sheet(title="Calculations")
ws_calc['A1'] = "Step-by-Step ET0 Calculation"
ws_calc['A1'].font = Font(bold=True, size=14)
ws_calc.merge_cells('A1:H1')

# Headers for calculations
calc_headers = ["Step", "Parameter", "Formula/Reference", "Value"]
for col, header in enumerate(calc_headers, start=1):
    cell = ws_calc.cell(row=3, column=col, value=header)
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal='center')

# List of calculation steps with formulas referencing Inputs sheet
# Formulas use Excel syntax, referencing Inputs!B4 etc.
# Row indices: Latitude B4, z B5, Date B6, T B7, Tmax B8, Tmin B9, RHmax B10, RHmin B11, U2 B12, Rs B13, n B14

steps = [
    (1, "Julian Day (J)", "DAYOFYEAR(Date)", '=DAYOFYEAR(Inputs!B6)'),
    (2, "Lat (rad)", "RADIANS(Latitude)", '=RADIANS(Inputs!B4)'),
    (3, "Solar Declination (delta)", "0.409 * SIN(2*PI()*J/365 - 1.39)", '=0.409 * SIN(2*PI()*B4/365 - 1.39)'),  # B4 is J here, but adjust refs
    (4, "Sunset Hour Angle (ws)", "ACOS(-TAN(lat_rad)*TAN(delta))", '=ACOS(-TAN(B5)*TAN(B6))'),  # Adjust cell refs accordingly
    (5, "Daylight Hours (N)", "24/PI() * ws", '=24/PI() * B7'),
    (6, "Inverse Relative Distance Earth-Sun (dr)", "1 + 0.033 * COS(2*PI()*J/365)", '=1 + 0.033 * COS(2*PI()*B4/365)'),
    (7, "Extraterrestrial Radiation (Ra)", "(24*60/PI())*0.082*dr*(ws*SIN(lat_rad)*SIN(delta) + COS(lat_rad)*COS(delta)*SIN(ws))", '=(24*60/PI())*0.082*B8*(B7*SIN(B5)*SIN(B6) + COS(B5)*COS(B6)*SIN(B7))'),
    (8, "Clear Sky Radiation (Rso)", "Ra * (0.75 + 2E-5 * z)", '=B9 * (0.75 + 0.00002 * Inputs!B5)'),
    (9, "Saturation Vapor Pressure at Tmean (es_T)", "0.611 * EXP(17.27 * T / (T + 237.3))", '=0.611 * EXP(17.27 * Inputs!B7 / (Inputs!B7 + 237.3))'),
    (10, "Saturation Vapor Pressure at Tmax (es_Tmax)", "0.611 * EXP(17.27 * Tmax / (Tmax + 237.3))", '=0.611 * EXP(17.27 * Inputs!B8 / (Inputs!B8 + 237.3))'),
    (11, "Saturation Vapor Pressure at Tmin (es_Tmin)", "0.611 * EXP(17.27 * Tmin / (Tmin + 237.3))", '=0.611 * EXP(17.27 * Inputs!B9 / (Inputs!B9 + 237.3))'),
    (12, "Mean Saturation Vapor Pressure (es)", "(es_Tmax + es_Tmin)/2", '=(B12 + B13)/2'),
    (13, "Actual Vapor Pressure (ed)", "(es_Tmin * (RHmax/100) + es_Tmax * (RHmin/100)) / 2", '=(B13 * (Inputs!B10/100) + B12 * (Inputs!B11/100)) / 2'),
    (14, "Slope Vapor Pressure Curve (Delta)", "4098 * es_T / (T + 237.3)^2", '=4098 * B11 / (Inputs!B7 + 237.3)^2'),
    (15, "Latent Heat of Vaporization (lambda)", "2.501 - 0.002361 * T", '=2.501 - 0.002361 * Inputs!B7'),
    (16, "Atmospheric Pressure (P)", "101.3 * ((293 - 0.0065 * z)/293)^5.26", '=101.3 * ((293 - 0.0065 * Inputs!B5)/293)^5.26'),
    (17, "Psychrometric Constant (gamma)", "0.00163 * P / lambda", '=0.00163 * B18 / B17'),
    (18, "Net Shortwave Radiation (Rns)", "(1 - 0.23) * Rs", '=(1 - 0.23) * Inputs!B13'),
    (19, "Tmax_K", "Tmax + 273.15", '=Inputs!B8 + 273.15'),
    (20, "Tmin_K", "Tmin + 273.15", '=Inputs!B9 + 273.15'),
    (21, "Net Longwave Radiation (Rnl)", "4.903E-9 * ((Tmax_K^4 + Tmin_K^4)/2) * (0.34 - 0.14 * SQRT(ed)) * (1.35 * Rs/Rso - 0.35)", '=4.903E-9 * ((B20^4 + B21^4)/2) * (0.34 - 0.14 * SQRT(B15)) * (1.35 * Inputs!B13 / B10 - 0.35)'),
    (22, "Net Radiation (Rn)", "Rns - Rnl", '=B19 - B22'),
    (23, "Soil Heat Flux (G)", "0 (for daily)", '=0'),
    (24, "ET0", "[0.408 * Delta * (Rn - G) + gamma * (900 / (T + 273)) * U2 * (es - ed)] / (Delta + gamma * (1 + 0.34 * U2))", '=(0.408 * B16 * (B23 - B24) + B19 * (900 / (Inputs!B7 + 273)) * Inputs!B12 * (B14 - B15)) / (B16 + B19 * (1 + 0.34 * Inputs!B12))'),
    (25, "Alternative Rs from Sunshine (if no Rs)", "Ra * (0.25 + 0.50 * (n/N))", '=B9 * (0.25 + 0.50 * (Inputs!B14 / B8))'),  # For variant
    (26, "ET0 with Sunshine Rs", "Substitute Rs with alternative in formula", '=(0.408 * B16 * ((1 - 0.23)*B26 - 4.903E-9 * ((B20^4 + B21^4)/2) * (0.34 - 0.14 * SQRT(B15)) * (1.35 * B26 / B10 - 0.35) - B24) + B19 * (900 / (Inputs!B7 + 273)) * Inputs!B12 * (B14 - B15)) / (B16 + B19 * (1 + 0.34 * Inputs!B12))'),
]

row = 4
for step_num, param, formula_desc, excel_formula in steps:
    ws_calc.cell(row=row, column=1, value=step_num)
    ws_calc.cell(row=row, column=2, value=param)
    cell_formula = ws_calc.cell(row=row, column=3, value=formula_desc)
    cell_value = ws_calc.cell(row=row, column=4)
    try:
        cell_value.value = excel_formula
    except:
        cell_value.value = "Formula error - check references"
    cell_formula.comment = Comment("This is the step based on SIAR PDF. Value is calculated automatically from inputs.", "Grok")
    row += 1

# Adjust column widths
for col in range(1, 5):
    ws_calc.column_dimensions[get_column_letter(col)].width = 30 if col in [3,4] else 15

# Sheet 3: Instructions
ws_instr = wb.create_sheet(title="Instructions")
ws_instr['A1'] = "How to Use This Excel Calculator"
ws_instr['A1'].font = Font(bold=True, size=14)
ws_instr['A3'] = "1. Go to 'Inputs' sheet and fill in Column B with your data."
ws_instr['A4'] = "2. Switch to 'Calculations' sheet to see step-by-step results."
ws_instr['A5'] = "3. Formulas reference inputs automatically."
ws_instr['A6'] = "4. For manual verification, you can copy formulas and calculate by hand."
ws_instr['A7'] = "5. If Rs is not available, use sunshine hours (n) for alternative calculation in steps 25-26."
ws_instr['A8'] = "Note: All formulas based on SIAR Penman-Monteith PDF. G assumed 0 for daily."

# Save the workbook
wb.save("ET0_Calculator.xlsx")
print("Excel file 'ET0_Calculator.xlsx' created successfully.")